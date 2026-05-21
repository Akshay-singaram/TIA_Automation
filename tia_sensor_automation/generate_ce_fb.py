"""
generate_ce_fb.py — Build Cause_n_Effect (FB8) with 200 interlocks from scratch.

Output routing is read from Process_CE.xlsx (row index = RS flip-flop index).
Rows 1-169 come from the Excel; RS[170..200] are Unassigned.
Blank rows (no X in any column) are also Unassigned.

Outputs: ESD, CSD_CS, Fire, Flammables, Ammonia_Release, CSD_HS, Unassigned
Static:  ESD_Acc1, ESD_Acc2  (ESD spans two CUs; combined in CU9)

One CompileUnit per output group (API requires each network to have one output):
  CU1 (ID=3)  — ESD[0:100]        → OR(100) → ESD_Acc1 static
  CU2 (ID=8)  — ESD[100:]         → OR(24)  → ESD_Acc2 static
  CU3 (ID=F)  — CSD_CS (26)       → OR(26)  → CSD_CS output
  CU4 (ID=14) — Unassigned (42)   → OR(42)  → Unassigned output
  CU5 (ID=19) — Fire (3)          → OR(3)   → Fire output
  CU6 (ID=1E) — Flammables (2)    → OR(2)   → Flammables output
  CU7 (ID=23) — Ammonia_Release(2)→ OR(2)   → Ammonia_Release output
  CU8 (ID=28) — CSD_HS (1)        → direct  → CSD_HS output
  CU9 (ID=2D) — ESD_Acc1|ESD_Acc2 → OR(2)  → ESD output

API import constraints:
  - OR Card per CompileUnit ≤ 100
  - OR.out → OR.in cascading rejected
  - Each CompileUnit must have exactly ONE output rung

Run as Administrator with TIA Portal open.
"""

import argparse
import os
import shutil
import sys
import tempfile

from config import EXPORT_DIR
from tia_portal import TIASession

COUNT     = 200
FB_NAME   = "Cause_n_Effect"
FB_NUMBER = 8
IFACE_NS  = "http://www.siemens.com/automation/Openness/SW/Interface/v5"
FLGNET_NS = "http://www.siemens.com/automation/Openness/SW/NetworkSource/FlgNet/v5"

PROCESS_CE_PATH = os.path.join(os.path.dirname(__file__), "Process_CE.xlsx")

# Maps raw Excel column headers → sanitised output variable names
_COL_SANITIZE = {
    "ESD":              "ESD",
    "Fire ":            "Fire",
    "Fire":             "Fire",
    "Flammables":       "Flammables",
    "Ammonia Release":  "Ammonia_Release",
    "CSD_HS":           "CSD_HS",
    "CSD_CS":           "CSD_CS",
}
UNASSIGNED = "Unassigned"

# Final output variables written to the FB Output section
OUTPUT_VARS = ["ESD", "CSD_CS", "Unassigned", "Fire", "Flammables", "Ammonia_Release", "CSD_HS"]
# Static accumulator variables (ESD split across two CUs)
STATIC_VARS = ["ESD_Acc1", "ESD_Acc2"]


# ---------------------------------------------------------------------------
# Excel reader
# ---------------------------------------------------------------------------

def read_excel_mapping() -> dict[str, list[int]]:
    """
    Return {output_name: [rs_indices]} for all 200 RS flip-flops.
    Row i (1-based) in the Excel → RS[i].
    RS[170..200] → Unassigned.
    """
    from collections import defaultdict
    from openpyxl import load_workbook

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name
    shutil.copy2(PROCESS_CE_PATH, tmp_path)

    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    finally:
        os.unlink(tmp_path)

    headers = rows[0]
    groups: dict[str, list[int]] = defaultdict(list)

    for row in rows[1:]:
        # Use the actual cause number from column 0, not the row position
        cause_raw = row[0] if row and len(row) > 0 else None
        if cause_raw is None or str(cause_raw).strip() == "":
            continue
        try:
            rs_idx = int(cause_raw)
        except (TypeError, ValueError):
            continue

        assigned = False
        for col_i, col_name in enumerate(headers[1:], 1):
            cell = row[col_i] if col_i < len(row) else None
            if cell is not None and str(cell).strip():
                raw = str(col_name) if col_name else ""
                out = _COL_SANITIZE.get(raw, raw.strip().replace(" ", "_"))
                groups[out].append(rs_idx)
                assigned = True
                break
        if not assigned:
            groups[UNASSIGNED].append(rs_idx)

    # Causes not in the Excel (e.g. 170..200) → Unassigned
    assigned_all = {idx for idxs in groups.values() for idx in idxs}
    for i in range(1, COUNT + 1):
        if i not in assigned_all:
            groups[UNASSIGNED].append(i)

    return dict(groups)


# ---------------------------------------------------------------------------
# Low-level XML helpers
# ---------------------------------------------------------------------------

def _enable_cause_name(i: int) -> str:
    return "Enable_Cause_1" if i == 1 else f"Enable_cause_{i}"


def _arr_access(uid: int, name: str, index: int) -> str:
    return (
        f'    <Access Scope="LocalVariable" UId="{uid}">\n'
        f'      <Symbol>\n'
        f'        <Component Name="{name}" AccessModifier="Array">\n'
        f'          <Access Scope="LiteralConstant">\n'
        f'            <Constant>\n'
        f'              <ConstantType>DInt</ConstantType>\n'
        f'              <ConstantValue>{index}</ConstantValue>\n'
        f'            </Constant>\n'
        f'          </Access>\n'
        f'        </Component>\n'
        f'      </Symbol>\n'
        f'    </Access>'
    )


def _var_access(uid: int, name: str) -> str:
    return (
        f'    <Access Scope="LocalVariable" UId="{uid}">\n'
        f'      <Symbol>\n'
        f'        <Component Name="{name}" />\n'
        f'      </Symbol>\n'
        f'    </Access>'
    )


def _flgnet(parts: list[str], wires: list[str]) -> str:
    return (
        f'<FlgNet xmlns="{FLGNET_NS}">\n'
        '  <Parts>\n' +
        '\n'.join(parts) + '\n'
        '  </Parts>\n'
        '  <Wires>\n' +
        '\n'.join(wires) + '\n'
        '  </Wires>\n'
        '</FlgNet>'
    )


# ---------------------------------------------------------------------------
# FlgNet builder — multiple RS groups in one CompileUnit
# ---------------------------------------------------------------------------

def _build_flgnet_grouped(
    groups: dict[str, list[int]],
    coil_vars: dict[str, str],
) -> str:
    """
    Build a FlgNet with RS-block logic for every RS index across all groups.

    TIA Portal Openness API requires strict flow order in the Parts/Wires sections:
      Parts:  ALL Access elements first (RS blocks in order, then output vars),
              then ALL Part elements (RS blocks in order, then OR+Coil per group).
      Wires:  Powerrail first, then for each RS block: 11 internal wires +
              RS.q→OR/Coil wire immediately, then OR→Coil wires for each group.

    Groups with 1 element wire RS.q → Coil.in directly (no OR block).
    Groups with >1 elements use OR(Card=N) where N = len(group).
    """
    groups    = {k: v for k, v in groups.items() if v}
    coil_vars = {k: v for k, v in coil_vars.items() if k in groups}

    uid = [0]

    def u() -> int:
        uid[0] += 1
        return uid[0]

    all_rs = sorted(set(idx for idxs in groups.values() for idx in idxs))

    # RS index → group name
    rs_to_group: dict[int, str] = {}
    for gname, idxs in groups.items():
        for i in idxs:
            rs_to_group[i] = gname

    # -----------------------------------------------------------------------
    # UID allocation — mirrors TIA Portal's export order exactly:
    #   Phase 1: Access UIDs for every RS block (6 per block)
    #   Phase 2: Access UIDs for output/accumulator variables (1 per group)
    #   Phase 3: Part UIDs for every RS block (6 per block)
    #   Phase 4: OR + Coil UIDs per group
    # -----------------------------------------------------------------------
    rs_acc: dict[int, dict] = {}
    for i in all_rs:
        rs_acc[i] = {
            "a_reset":  u(), "a_enable": u(), "a_bypass": u(),
            "a_en_st":  u(), "a_cause":  u(), "a_latch":  u(),
        }

    grp_acc: dict[str, int] = {gname: u() for gname in groups}

    rs_prt: dict[int, dict] = {}
    for i in all_rs:
        rs_prt[i] = {
            "p_c_reset":  u(), "p_c_enable": u(), "p_c_bypass": u(),
            "p_coil_en":  u(), "p_c_cause":  u(), "p_rs":       u(),
        }

    grp_prt: dict[str, dict] = {}
    for gname, idxs in groups.items():
        use_or = len(idxs) > 1
        grp_prt[gname] = {
            "use_or": use_or,
            "card":   len(idxs),
            "p_or":   u() if use_or else None,
            "p_coil": u(),
        }

    # -----------------------------------------------------------------------
    # Parts — all Accesses, then all Parts (flow order)
    # -----------------------------------------------------------------------
    parts: list[str] = []

    for i in all_rs:
        a = rs_acc[i]
        parts += [
            _arr_access(a["a_reset"],  "InterlockReset",        i),
            _var_access(a["a_enable"], _enable_cause_name(i)),
            _arr_access(a["a_bypass"], "InterlockBypass",       i),
            _arr_access(a["a_en_st"],  "InterlockEnableStatus", i),
            _var_access(a["a_cause"],  f"Cause_{i}"),
            _arr_access(a["a_latch"],  "InterlockLatchStatus",  i),
        ]

    for gname, a_uid in grp_acc.items():
        parts.append(_var_access(a_uid, coil_vars[gname]))

    for i in all_rs:
        p = rs_prt[i]
        parts += [
            f'    <Part Name="Contact" UId="{p["p_c_reset"]}" />',
            f'    <Part Name="Contact" UId="{p["p_c_enable"]}" />',
            f'    <Part Name="Contact" UId="{p["p_c_bypass"]}">\n'
            f'      <Negated Name="operand" />\n'
            f'    </Part>',
            f'    <Part Name="Coil" UId="{p["p_coil_en"]}" />',
            f'    <Part Name="Contact" UId="{p["p_c_cause"]}" />',
            f'    <Part Name="Rs" UId="{p["p_rs"]}" />',
        ]

    for gname, gp in grp_prt.items():
        if gp["use_or"]:
            parts.append(
                f'    <Part Name="O" UId="{gp["p_or"]}">\n'
                f'      <TemplateValue Name="Card" Type="Cardinality">{gp["card"]}</TemplateValue>\n'
                f'    </Part>'
            )
        parts.append(f'    <Part Name="Coil" UId="{gp["p_coil"]}" />')

    # -----------------------------------------------------------------------
    # Wires — powerrail first, then per-RS-block (11 + q wire), then OR→Coil
    # -----------------------------------------------------------------------
    wires: list[str] = []

    pr_lines = [f'    <Wire UId="{u()}">', '      <Powerrail />']
    for i in all_rs:
        p = rs_prt[i]
        pr_lines.append(f'      <NameCon UId="{p["p_c_reset"]}" Name="in" />')
        pr_lines.append(f'      <NameCon UId="{p["p_c_enable"]}" Name="in" />')
    pr_lines.append('    </Wire>')
    wires.append('\n'.join(pr_lines))

    pin_count: dict[str, int] = {gname: 0 for gname in groups}

    for i in all_rs:
        a = rs_acc[i]
        p = rs_prt[i]
        gname = rs_to_group[i]
        gp = grp_prt[gname]

        wires += [
            f'    <Wire UId="{u()}">\n'
            f'      <IdentCon UId="{a["a_reset"]}" />\n'
            f'      <NameCon UId="{p["p_c_reset"]}" Name="operand" />\n'
            f'    </Wire>',

            f'    <Wire UId="{u()}">\n'
            f'      <NameCon UId="{p["p_c_reset"]}" Name="out" />\n'
            f'      <NameCon UId="{p["p_rs"]}" Name="r" />\n'
            f'    </Wire>',

            f'    <Wire UId="{u()}">\n'
            f'      <IdentCon UId="{a["a_enable"]}" />\n'
            f'      <NameCon UId="{p["p_c_enable"]}" Name="operand" />\n'
            f'    </Wire>',

            f'    <Wire UId="{u()}">\n'
            f'      <NameCon UId="{p["p_c_enable"]}" Name="out" />\n'
            f'      <NameCon UId="{p["p_c_bypass"]}" Name="in" />\n'
            f'    </Wire>',

            f'    <Wire UId="{u()}">\n'
            f'      <IdentCon UId="{a["a_bypass"]}" />\n'
            f'      <NameCon UId="{p["p_c_bypass"]}" Name="operand" />\n'
            f'    </Wire>',

            f'    <Wire UId="{u()}">\n'
            f'      <NameCon UId="{p["p_c_bypass"]}" Name="out" />\n'
            f'      <NameCon UId="{p["p_coil_en"]}" Name="in" />\n'
            f'    </Wire>',

            f'    <Wire UId="{u()}">\n'
            f'      <IdentCon UId="{a["a_en_st"]}" />\n'
            f'      <NameCon UId="{p["p_coil_en"]}" Name="operand" />\n'
            f'    </Wire>',

            f'    <Wire UId="{u()}">\n'
            f'      <NameCon UId="{p["p_coil_en"]}" Name="out" />\n'
            f'      <NameCon UId="{p["p_c_cause"]}" Name="in" />\n'
            f'    </Wire>',

            f'    <Wire UId="{u()}">\n'
            f'      <IdentCon UId="{a["a_cause"]}" />\n'
            f'      <NameCon UId="{p["p_c_cause"]}" Name="operand" />\n'
            f'    </Wire>',

            f'    <Wire UId="{u()}">\n'
            f'      <NameCon UId="{p["p_c_cause"]}" Name="out" />\n'
            f'      <NameCon UId="{p["p_rs"]}" Name="s1" />\n'
            f'    </Wire>',

            f'    <Wire UId="{u()}">\n'
            f'      <IdentCon UId="{a["a_latch"]}" />\n'
            f'      <NameCon UId="{p["p_rs"]}" Name="operand" />\n'
            f'    </Wire>',
        ]

        # RS.q → OR.inN (or directly to Coil for single-element groups)
        pin_count[gname] += 1
        if gp["use_or"]:
            wires.append(
                f'    <Wire UId="{u()}">\n'
                f'      <NameCon UId="{p["p_rs"]}" Name="q" />\n'
                f'      <NameCon UId="{gp["p_or"]}" Name="in{pin_count[gname]}" />\n'
                f'    </Wire>'
            )
        else:
            wires.append(
                f'    <Wire UId="{u()}">\n'
                f'      <NameCon UId="{p["p_rs"]}" Name="q" />\n'
                f'      <NameCon UId="{gp["p_coil"]}" Name="in" />\n'
                f'    </Wire>'
            )

    for gname, gp in grp_prt.items():
        if gp["use_or"]:
            wires.append(
                f'    <Wire UId="{u()}">\n'
                f'      <NameCon UId="{gp["p_or"]}" Name="out" />\n'
                f'      <NameCon UId="{gp["p_coil"]}" Name="in" />\n'
                f'    </Wire>'
            )
        wires.append(
            f'    <Wire UId="{u()}">\n'
            f'      <IdentCon UId="{grp_acc[gname]}" />\n'
            f'      <NameCon UId="{gp["p_coil"]}" Name="operand" />\n'
            f'    </Wire>'
        )

    return _flgnet(parts, wires)


def _build_flgnet_combine(acc_vars: list[str], output_var: str) -> str:
    """
    CU4: combines ESD_Acc1 + ESD_Acc2 via contacts → OR(2) → ESD output coil.
    acc_vars must have exactly 2 entries.
    """
    uid = [0]

    def u() -> int:
        uid[0] += 1
        return uid[0]

    card = len(acc_vars)

    # Allocate access UIDs
    a_accs  = [u() for _ in acc_vars]
    a_out   = u()
    # Part UIDs
    p_conts = [u() for _ in acc_vars]
    p_or    = u()
    p_coil  = u()
    # Wire UIDs
    w_pr    = u()
    w_ops   = [u() for _ in acc_vars]
    w_outs  = [u() for _ in acc_vars]
    w_or    = u()
    w_op    = u()

    parts = []
    for a_acc, acc_var in zip(a_accs, acc_vars):
        parts.append(_var_access(a_acc, acc_var))
    parts.append(_var_access(a_out, output_var))
    for pc in p_conts:
        parts.append(f'    <Part Name="Contact" UId="{pc}" />')
    parts += [
        f'    <Part Name="O" UId="{p_or}">\n'
        f'      <TemplateValue Name="Card" Type="Cardinality">{card}</TemplateValue>\n'
        f'    </Part>',
        f'    <Part Name="Coil" UId="{p_coil}" />',
    ]

    # Powerrail → all contacts
    pr_lines = [f'    <Wire UId="{w_pr}">', '      <Powerrail />']
    for pc in p_conts:
        pr_lines.append(f'      <NameCon UId="{pc}" Name="in" />')
    pr_lines.append('    </Wire>')

    wires = ['\n'.join(pr_lines)]
    for j, (a_acc, pc, w_op_j, w_out_j) in enumerate(
        zip(a_accs, p_conts, w_ops, w_outs), 1
    ):
        wires += [
            f'    <Wire UId="{w_op_j}">\n'
            f'      <IdentCon UId="{a_acc}" />\n'
            f'      <NameCon UId="{pc}" Name="operand" />\n'
            f'    </Wire>',
            f'    <Wire UId="{w_out_j}">\n'
            f'      <NameCon UId="{pc}" Name="out" />\n'
            f'      <NameCon UId="{p_or}" Name="in{j}" />\n'
            f'    </Wire>',
        ]
    wires += [
        f'    <Wire UId="{w_or}">\n'
        f'      <NameCon UId="{p_or}" Name="out" />\n'
        f'      <NameCon UId="{p_coil}" Name="in" />\n'
        f'    </Wire>',
        f'    <Wire UId="{w_op}">\n'
        f'      <IdentCon UId="{a_out}" />\n'
        f'      <NameCon UId="{p_coil}" Name="operand" />\n'
        f'    </Wire>',
    ]

    return _flgnet(parts, wires)


# ---------------------------------------------------------------------------
# Interface builder
# ---------------------------------------------------------------------------

def _build_interface() -> str:
    lines = [f'<Sections xmlns="{IFACE_NS}">', '  <Section Name="Input">']
    for i in range(1, COUNT + 1):
        lines.append(f'    <Member Name="Cause_{i}" Datatype="Bool" />')
    lines.append(f'    <Member Name="InterlockReset" Datatype="Array[1..{COUNT}] of Bool" />')
    lines.append(f'    <Member Name="InterlockBypass" Datatype="Array[1..{COUNT}] of Bool" />')
    for i in range(1, COUNT + 1):
        lines.append(f'    <Member Name="{_enable_cause_name(i)}" Datatype="Bool" />')
    lines += [
        '  </Section>',
        '  <Section Name="Output">',
    ]
    for out_var in OUTPUT_VARS:
        lines.append(f'    <Member Name="{out_var}" Datatype="Bool" />')
    lines += [
        f'    <Member Name="InterlockLatchStatus" Datatype="Array[1..{COUNT}] of Bool" />',
        f'    <Member Name="InterlockEnableStatus" Datatype="Array[1..{COUNT}] of Bool" />',
        '  </Section>',
        '  <Section Name="InOut" />',
        '  <Section Name="Static">',
    ]
    for sv in STATIC_VARS:
        lines.append(f'    <Member Name="{sv}" Datatype="Bool" />')
    lines += [
        '  </Section>',
        '  <Section Name="Temp" />',
        '  <Section Name="Constant" />',
        '</Sections>',
    ]
    return '\n'.join(lines)


# ---------------------------------------------------------------------------
# Full XML assembler
# ---------------------------------------------------------------------------

def _compile_unit(cu_id: str, flgnet: str,
                  comment_id: str, comment_item_id: str,
                  title_id: str,   title_item_id: str) -> str:
    return (
        f'      <SW.Blocks.CompileUnit ID="{cu_id}" CompositionName="CompileUnits">\n'
        f'        <AttributeList>\n'
        f'          <NetworkSource>{flgnet}</NetworkSource>\n'
        f'          <ProgrammingLanguage>LAD</ProgrammingLanguage>\n'
        f'        </AttributeList>\n'
        f'        <ObjectList>\n'
        f'          <MultilingualText ID="{comment_id}" CompositionName="Comment">\n'
        f'            <ObjectList>\n'
        f'              <MultilingualTextItem ID="{comment_item_id}" CompositionName="Items">\n'
        f'                <AttributeList><Culture>en-US</Culture><Text /></AttributeList>\n'
        f'              </MultilingualTextItem>\n'
        f'            </ObjectList>\n'
        f'          </MultilingualText>\n'
        f'          <MultilingualText ID="{title_id}" CompositionName="Title">\n'
        f'            <ObjectList>\n'
        f'              <MultilingualTextItem ID="{title_item_id}" CompositionName="Items">\n'
        f'                <AttributeList><Culture>en-US</Culture><Text /></AttributeList>\n'
        f'              </MultilingualTextItem>\n'
        f'            </ObjectList>\n'
        f'          </MultilingualText>\n'
        f'        </ObjectList>\n'
        f'      </SW.Blocks.CompileUnit>\n'
    )


def _build_full_xml() -> str:
    groups = read_excel_mapping()

    esd        = groups.get("ESD",             [])
    csdcs      = groups.get("CSD_CS",          [])
    unassigned = groups.get("Unassigned",      [])
    fire       = groups.get("Fire",            [])
    flam       = groups.get("Flammables",      [])
    nh3        = groups.get("Ammonia_Release", [])
    hs         = groups.get("CSD_HS",          [])

    print("  Output distribution:")
    for name, lst in [("ESD", esd), ("CSD_CS", csdcs), ("Unassigned", unassigned),
                      ("Fire", fire), ("Flammables", flam),
                      ("Ammonia_Release", nh3), ("CSD_HS", hs)]:
        print(f"    {name}: {len(lst)} causes")

    # One CompileUnit per output group — API requires single output rung per network.
    flgnet_cu1 = _build_flgnet_grouped({"ESD":             esd[:100]},  {"ESD":             "ESD_Acc1"})
    flgnet_cu2 = _build_flgnet_grouped({"ESD":             esd[100:]},  {"ESD":             "ESD_Acc2"})
    flgnet_cu3 = _build_flgnet_grouped({"CSD_CS":          csdcs},      {"CSD_CS":          "CSD_CS"})
    flgnet_cu4 = _build_flgnet_grouped({"Unassigned":      unassigned}, {"Unassigned":      "Unassigned"})
    flgnet_cu5 = _build_flgnet_grouped({"Fire":            fire},       {"Fire":            "Fire"})
    flgnet_cu6 = _build_flgnet_grouped({"Flammables":      flam},       {"Flammables":      "Flammables"})
    flgnet_cu7 = _build_flgnet_grouped({"Ammonia_Release": nh3},        {"Ammonia_Release": "Ammonia_Release"})
    flgnet_cu8 = _build_flgnet_grouped({"CSD_HS":          hs},         {"CSD_HS":          "CSD_HS"})
    flgnet_cu9 = _build_flgnet_combine(["ESD_Acc1", "ESD_Acc2"],        "ESD")

    cu1 = _compile_unit("3",  flgnet_cu1, "4",  "5",  "6",  "7")
    cu2 = _compile_unit("8",  flgnet_cu2, "9",  "A",  "B",  "C")
    cu3 = _compile_unit("F",  flgnet_cu3, "10", "11", "12", "13")
    cu4 = _compile_unit("14", flgnet_cu4, "15", "16", "17", "18")
    cu5 = _compile_unit("19", flgnet_cu5, "1A", "1B", "1C", "1D")
    cu6 = _compile_unit("1E", flgnet_cu6, "1F", "20", "21", "22")
    cu7 = _compile_unit("23", flgnet_cu7, "24", "25", "26", "27")
    cu8 = _compile_unit("28", flgnet_cu8, "29", "2A", "2B", "2C")
    cu9 = _compile_unit("2D", flgnet_cu9, "2E", "2F", "30", "31")

    interface = _build_interface()

    return (
        '<?xml version="1.0" encoding="utf-8"?>\n'
        '<Document>\n'
        '  <Engineering version="V19" />\n'
        '  <DocumentInfo>\n'
        '    <ExportSetting>WithDefaults, RecursiveWithDefaults</ExportSetting>\n'
        '  </DocumentInfo>\n'
        '  <SW.Blocks.FB ID="0">\n'
        '    <AttributeList>\n'
        f'      <Interface>{interface}</Interface>\n'
        '      <MemoryLayout>Optimized</MemoryLayout>\n'
        '      <MemoryReserve>100</MemoryReserve>\n'
        f'      <Name>{FB_NAME}</Name>\n'
        '      <Namespace />\n'
        f'      <Number>{FB_NUMBER}</Number>\n'
        '      <ProgrammingLanguage>LAD</ProgrammingLanguage>\n'
        '      <SetENOAutomatically>false</SetENOAutomatically>\n'
        '    </AttributeList>\n'
        '    <ObjectList>\n'
        '      <MultilingualText ID="1" CompositionName="Comment">\n'
        '        <ObjectList>\n'
        '          <MultilingualTextItem ID="2" CompositionName="Items">\n'
        '            <AttributeList><Culture>en-US</Culture><Text /></AttributeList>\n'
        '          </MultilingualTextItem>\n'
        '        </ObjectList>\n'
        '      </MultilingualText>\n'
        + cu1 + cu2 + cu3 + cu4 + cu5 + cu6 + cu7 + cu8 + cu9 +
        '      <MultilingualText ID="D" CompositionName="Title">\n'
        '        <ObjectList>\n'
        '          <MultilingualTextItem ID="E" CompositionName="Items">\n'
        '            <AttributeList><Culture>en-US</Culture><Text /></AttributeList>\n'
        '          </MultilingualTextItem>\n'
        '        </ObjectList>\n'
        '      </MultilingualText>\n'
        '    </ObjectList>\n'
        '  </SW.Blocks.FB>\n'
        '</Document>'
    )


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> int:
    global FB_NAME, FB_NUMBER, PROCESS_CE_PATH

    parser = argparse.ArgumentParser(
        description="Build a Cause-and-Effect FB from an Excel mapping file."
    )
    parser.add_argument(
        "--excel", default=None,
        help="Path to the Process_CE Excel file (default: Process_CE.xlsx next to this script)"
    )
    parser.add_argument(
        "--fb-name", default=FB_NAME,
        help=f"Name of the FB to create/replace in TIA Portal (default: {FB_NAME})"
    )
    parser.add_argument(
        "--fb-number", type=int, default=FB_NUMBER,
        help=f"Block number to assign to the FB (default: {FB_NUMBER})"
    )
    args = parser.parse_args()

    FB_NAME   = args.fb_name
    FB_NUMBER = args.fb_number
    if args.excel:
        PROCESS_CE_PATH = os.path.abspath(args.excel)

    print(f"Building '{FB_NAME}' (FB{FB_NUMBER}) with {COUNT} interlocks...")
    print(f"  Excel: {PROCESS_CE_PATH}")

    with TIASession() as session:
        print("\n[1] Generating XML from scratch...")
        xml_path = os.path.join(EXPORT_DIR, f"{FB_NAME}.xml")
        os.makedirs(EXPORT_DIR, exist_ok=True)
        with open(xml_path, 'w', encoding='utf-8') as f:
            f.write(_build_full_xml())
        print(f"  Written: {xml_path}")

        print("\n[2] Importing FB (deleting existing if present)...")
        session.import_fc(xml_path, FB_NAME)

        print("\n[3] Compiling...")
        result = session.compile()
        if result.ErrorCount > 0:
            print(f"  WARNING: {result.ErrorCount} compile error(s).")

    print("\nDone.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
