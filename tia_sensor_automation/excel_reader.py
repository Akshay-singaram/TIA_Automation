from openpyxl import load_workbook

from config import (
    COL_BLOCK_GROUP,
    COL_SENSOR_NAME,
    COL_SOURCE_FB,
    COL_TARGET_FC,
    EXCEL_FILE_PATH,
    COL_DB_NAME,
    COL_ARRAY_NAME,
    COL_ARRAY_INDEX,
    COL_VARIABLE_NAME,
    COL_DEFAULT_VALUE,
    DB_DEFAULTS_EXCEL_PATH,
)

REQUIRED_COLUMNS = [COL_SENSOR_NAME, COL_SOURCE_FB, COL_TARGET_FC, COL_BLOCK_GROUP]


def read_sensor_rows(file_path: str | None = None) -> list[dict]:
    """
    Read sensor/actuator rows from the first sheet of an Excel workbook.

    Required columns (set in config.py):
      COL_SENSOR_NAME  — name of the sensor or actuator
      COL_SOURCE_FB    — FB to instantiate for this row
      COL_TARGET_FC    — LAD FC to inject the network into
      COL_BLOCK_GROUP  — block group path for the instance DB

    Rows where Sensor_Name is blank are silently skipped.
    Returns a list of dicts with keys: sensor_name, source_fb, target_fc, block_group.
    """
    path = file_path or EXCEL_FILE_PATH

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"Workbook '{path}' has no active sheet.")

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ValueError(f"Workbook '{path}' appears to be empty.")

    headers = [str(h).strip() if h is not None else "" for h in header_row]

    missing = [c for c in REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ValueError(
            f"Missing column(s) in '{path}': {missing}\n"
            f"  Headers found: {headers}"
        )

    idx = {col: headers.index(col) for col in REQUIRED_COLUMNS}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        sensor = row[idx[COL_SENSOR_NAME]] if idx[COL_SENSOR_NAME] < len(row) else None
        if sensor is None or not str(sensor).strip():
            continue
        rows.append({
            "sensor_name": str(sensor).strip(),
            "source_fb":   str(row[idx[COL_SOURCE_FB]]).strip(),
            "target_fc":   str(row[idx[COL_TARGET_FC]]).strip(),
            "block_group": str(row[idx[COL_BLOCK_GROUP]]).strip(),
        })

    wb.close()
    return rows


DB_DEFAULTS_REQUIRED_COLUMNS = [
    COL_DB_NAME, COL_ARRAY_NAME, COL_ARRAY_INDEX, COL_VARIABLE_NAME, COL_DEFAULT_VALUE
]


def read_db_default_rows(file_path: str | None = None) -> list[dict]:
    """
    Read DB default-value update rows from db_defaults.xlsx.

    Required columns:
      DB_Name, Array_Name, Array_Index, Variable_Name, Default_Value

    Rows where DB_Name is blank are skipped.
    Returns a list of dicts with keys:
      db_name, array_name, array_index (int), variable_name, default_value (str)
    """
    path = file_path or DB_DEFAULTS_EXCEL_PATH

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ValueError(f"Workbook '{path}' has no active sheet.")

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        raise ValueError(f"Workbook '{path}' appears to be empty.")

    headers = [str(h).strip() if h is not None else "" for h in header_row]

    missing = [c for c in DB_DEFAULTS_REQUIRED_COLUMNS if c not in headers]
    if missing:
        raise ValueError(
            f"Missing column(s) in '{path}': {missing}\n"
            f"  Headers found: {headers}"
        )

    idx = {col: headers.index(col) for col in DB_DEFAULTS_REQUIRED_COLUMNS}

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        db_name = row[idx[COL_DB_NAME]] if idx[COL_DB_NAME] < len(row) else None
        if db_name is None or not str(db_name).strip():
            continue
        try:
            array_index = int(str(row[idx[COL_ARRAY_INDEX]]))
        except (TypeError, ValueError):
            raise ValueError(
                f"Array_Index must be an integer, got: {row[idx[COL_ARRAY_INDEX]]!r}"
            )
        rows.append({
            "db_name":       str(db_name).strip(),
            "array_name":    str(row[idx[COL_ARRAY_NAME]]).strip(),
            "array_index":   array_index,
            "variable_name": str(row[idx[COL_VARIABLE_NAME]]).strip(),
            "default_value": str(row[idx[COL_DEFAULT_VALUE]]).strip(),
        })

    wb.close()
    return rows
